"""QA Excel ハンドラー

RFP質疑一覧のExcel操作を行うユーティリティ。
- Boxから取得したQAテンプレートExcelのカラム構造を読み取る
- qa_draft.jsonをテンプレートフォーマットに合わせてExcelに書き出す
- 顧客が回答済みのExcel/Markdownを取り込み、answers.mdを生成する

使い方:
    # テンプレートを使ってExcel出力
    python3 tools/qa_excel_handler.py write-excel \\
        --template path/to/template.xlsx \\
        --qa-json  path/to/qa_draft.json \\
        --output   path/to/qa_list.xlsx

    # 回答済みExcelを取り込む
    python3 tools/qa_excel_handler.py import-answers \\
        --answered path/to/qa_answered.xlsx \\
        --qa-json  path/to/qa_draft.json \\
        --output   path/to/answers.md

    # 回答済みMarkdownを取り込む
    python3 tools/qa_excel_handler.py import-answers \\
        --answered path/to/qa_answered.md \\
        --qa-json  path/to/qa_draft.json \\
        --output   path/to/answers.md

関連Skill:
    .claude/skills/rfp-qa-generator/SKILL.md
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any


# カラムマッピング定義
# キー: 内部フィールド名, 値: マッピング対象キーワードのリスト（部分一致・大小無視）
COLUMN_KEYWORD_MAP: dict[str, list[str]] = {
    "id": ["no", "番号", "id"],
    "perspective": ["観点", "カテゴリ", "区分"],
    "priority": ["優先", "重要度", "priority"],
    "question": ["質疑", "質問", "question", "内容"],
    "background": ["背景", "目的", "理由"],
    "reference": ["参照", "出典", "箇所", "reference"],
    "answer": ["回答", "answer", "顧客回答"],
}

# 回答取下げキーワード
WITHDRAW_KEYWORDS: list[str] = ["取下げ", "取り下げ", "不要", "キャンセル", "cancel"]


def _load_qa_json(qa_json_path: Path) -> dict[str, Any]:
    """qa_draft.jsonを読み込む。

    Args:
        qa_json_path: qa_draft.jsonのパス。

    Returns:
        qa_draft.jsonの内容を表す辞書。

    Raises:
        SystemExit: ファイルが存在しない、またはJSONパースエラーの場合。
    """
    if not qa_json_path.exists():
        print(f"エラー: {qa_json_path} が見つかりません", file=sys.stderr)
        sys.exit(1)
    try:
        return json.loads(qa_json_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as e:
        print(f"エラー: {qa_json_path} のJSONパースに失敗しました: {e}", file=sys.stderr)
        sys.exit(1)


def _detect_column_mapping(headers: list[str]) -> dict[str, int]:
    """ヘッダー行のカラム名から内部フィールドとのマッピングを生成する。

    各ヘッダー文字列に対して、COLUMN_KEYWORD_MAPのキーワードを部分一致（大小無視）で
    検索し、最初にマッチしたフィールドに割り当てる。

    Args:
        headers: Excelヘッダー行のカラム名リスト（インデックス順）。

    Returns:
        内部フィールド名 → Excelカラムインデックス のマッピング辞書。
        マッピングできたフィールドのみを含む。
    """
    mapping: dict[str, int] = {}
    for col_idx, header in enumerate(headers):
        header_lower = header.lower().strip()
        for field, keywords in COLUMN_KEYWORD_MAP.items():
            if field in mapping:
                continue
            if any(kw.lower() in header_lower for kw in keywords):
                mapping[field] = col_idx
                break
    return mapping


def _qa_item_to_row(
    item: dict[str, Any],
    mapping: dict[str, int],
    total_cols: int,
) -> list[Any]:
    """QAアイテムをExcelの1行データに変換する。

    Args:
        item: qa_draft.jsonのqa_list要素。
        mapping: 内部フィールド名 → Excelカラムインデックス のマッピング。
        total_cols: Excelの総カラム数（マッピング外カラムは空文字で埋める）。

    Returns:
        Excelの1行分のデータリスト。
    """
    row: list[Any] = [""] * total_cols
    for field, col_idx in mapping.items():
        value = item.get(field)
        if value is None:
            value = ""
        # answerはまだ空（未回答）の場合は「（未回答）」と表示
        if field == "answer" and not value:
            value = "（未回答）"
        row[col_idx] = str(value)
    return row


def cmd_write_excel(
    template_path: Path,
    qa_json_path: Path,
    output_path: Path,
) -> None:
    """Boxテンプレートを使ってQA一覧をExcelに出力する。

    テンプレートExcelのヘッダー行を自動検出し、内部フィールドとマッピングして
    qa_draft.jsonの質疑一覧を書き込む。

    Args:
        template_path: BoxからDLしたQAテンプレートExcelのパス。
        qa_json_path: qa_draft.jsonのパス。
        output_path: 出力先Excelファイルのパス。

    Raises:
        SystemExit: openpyxlが未インストール、またはファイルエラーの場合。
    """
    try:
        import openpyxl
    except ImportError:
        print("エラー: openpyxlが未インストールです。pip install openpyxl を実行してください", file=sys.stderr)
        sys.exit(1)

    if not template_path.exists():
        print(f"エラー: テンプレートファイル {template_path} が見つかりません", file=sys.stderr)
        sys.exit(1)

    qa_data = _load_qa_json(qa_json_path)
    qa_list: list[dict[str, Any]] = qa_data.get("qa_list", [])

    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # ヘッダー行を検出（最初に文字列が含まれる行をヘッダーとみなす）
    header_row_idx: int | None = None
    headers: list[str] = []
    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        row_values = [str(cell.value or "").strip() for cell in row]
        non_empty = [v for v in row_values if v]
        if len(non_empty) >= 2:  # 2列以上に値があればヘッダー行とみなす
            header_row_idx = row_idx
            headers = row_values
            break

    if header_row_idx is None:
        print("エラー: テンプレートのヘッダー行を検出できませんでした", file=sys.stderr)
        sys.exit(1)

    mapping = _detect_column_mapping(headers)
    if not mapping:
        print("警告: カラムマッピングが1件も成立しませんでした。テンプレートのヘッダー名を確認してください", file=sys.stderr)

    print(f"カラムマッピング: {mapping}")
    total_cols = len(headers)

    # テンプレートのデータ行（ヘッダー行以降）をクリア
    max_row = ws.max_row
    if max_row > header_row_idx:
        for row in ws.iter_rows(min_row=header_row_idx + 1, max_row=max_row):
            for cell in row:
                cell.value = None

    # QAデータを書き込む
    for item in qa_list:
        row_data = _qa_item_to_row(item, mapping, total_cols)
        ws.append(row_data)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Excel出力完了: {output_path} ({len(qa_list)}件)")


def _parse_answered_excel(answered_path: Path) -> dict[str, str]:
    """回答済みExcelを解析し、質疑IDと回答のマッピングを返す。

    Args:
        answered_path: 顧客が回答済みのExcelファイルのパス。

    Returns:
        質疑ID（例: "Q001"）→ 回答文字列 のマッピング辞書。
        IDが取得できない行はスキップする。

    Raises:
        SystemExit: openpyxlが未インストール、またはファイルエラーの場合。
    """
    try:
        import openpyxl
    except ImportError:
        print("エラー: openpyxlが未インストールです。pip install openpyxl を実行してください", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(answered_path, read_only=True)
    ws = wb.active

    # ヘッダー行の検出とマッピング
    header_row_idx: int | None = None
    mapping: dict[str, int] = {}
    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        row_values = [str(cell.value or "").strip() for cell in row]
        non_empty = [v for v in row_values if v]
        if len(non_empty) >= 2:
            header_row_idx = row_idx
            mapping = _detect_column_mapping(row_values)
            break

    if header_row_idx is None or "id" not in mapping or "answer" not in mapping:
        print("警告: 回答済みExcelからIDまたは回答カラムを検出できませんでした", file=sys.stderr)
        wb.close()
        return {}

    id_col = mapping["id"]
    answer_col = mapping["answer"]

    answers: dict[str, str] = {}
    for row in ws.iter_rows(min_row=header_row_idx + 1):
        row_values = [str(cell.value or "").strip() for cell in row]
        qa_id = row_values[id_col] if id_col < len(row_values) else ""
        answer = row_values[answer_col] if answer_col < len(row_values) else ""
        if qa_id and answer:
            answers[qa_id] = answer

    wb.close()
    return answers


def _parse_answered_markdown(answered_path: Path) -> dict[str, str]:
    """回答済みMarkdownを解析し、質疑IDと回答のマッピングを返す。

    qa_list.mdのMarkdown表から「顧客回答」列を読み取る。
    「（未回答）」のセルはスキップする。

    Args:
        answered_path: 顧客が回答済みのMarkdownファイルのパス。

    Returns:
        質疑ID（例: "Q001"）→ 回答文字列 のマッピング辞書。
    """
    text = answered_path.read_text(encoding="utf-8")
    answers: dict[str, str] = {}

    # Markdown表の行を解析（| Q001 | ... | 回答 | の形式）
    # ヘッダー行でAnswerカラムのインデックスを特定する
    answer_col_idx: int | None = None
    id_col_idx: int | None = None

    for line in text.splitlines():
        line = line.strip()
        if not line.startswith("|"):
            continue
        cols = [c.strip() for c in line.strip("|").split("|")]

        # ヘッダー行の検出
        if answer_col_idx is None:
            for i, col in enumerate(cols):
                if any(kw.lower() in col.lower() for kw in COLUMN_KEYWORD_MAP["answer"]):
                    answer_col_idx = i
                if any(kw.lower() in col.lower() for kw in COLUMN_KEYWORD_MAP["id"]):
                    id_col_idx = i
            continue

        # セパレーター行はスキップ
        if all(re.match(r"^[-:]+$", c) for c in cols if c):
            continue

        if id_col_idx is None or answer_col_idx is None:
            continue
        if id_col_idx >= len(cols) or answer_col_idx >= len(cols):
            continue

        qa_id = cols[id_col_idx]
        answer = cols[answer_col_idx]

        if qa_id and answer and answer != "（未回答）":
            answers[qa_id] = answer

    return answers


def _generate_reflection_note(qa_item: dict[str, Any], answer: str) -> str:
    """回答内容から提案への示唆を生成する。

    Args:
        qa_item: qa_draft.jsonのqa_list要素。
        answer: 顧客の回答文字列。

    Returns:
        提案への示唆を表す文字列（簡易的な要約）。
    """
    perspective = qa_item.get("perspective", "")
    priority = qa_item.get("priority", "")

    if len(answer) > 100:
        summary = answer[:97] + "..."
    else:
        summary = answer

    if priority == "必須":
        return f"[必須反映] {perspective}: {summary}"
    return f"[推奨反映] {perspective}: {summary}"


def cmd_import_answers(
    answered_path: Path,
    qa_json_path: Path,
    output_path: Path,
) -> None:
    """回答済みExcel/Markdownを取り込み、answers.mdを生成する。

    qa_draft.jsonに回答内容を反映し、answers.mdに整理されたサマリを出力する。
    qa_draft.jsonも更新する（answerとstatusフィールドを設定）。

    Args:
        answered_path: 顧客が回答済みのExcelまたはMarkdownファイルのパス。
        qa_json_path: qa_draft.jsonのパス。
        output_path: answers.md の出力先パス。

    Raises:
        SystemExit: ファイルエラーの場合。
    """
    if not answered_path.exists():
        print(f"エラー: {answered_path} が見つかりません", file=sys.stderr)
        sys.exit(1)

    # 回答の解析（ファイル形式で分岐）
    suffix = answered_path.suffix.lower()
    if suffix in (".xlsx", ".xls"):
        answers = _parse_answered_excel(answered_path)
    elif suffix in (".md", ".markdown", ".txt"):
        answers = _parse_answered_markdown(answered_path)
    else:
        print(f"エラー: 未対応のファイル形式です: {suffix}", file=sys.stderr)
        sys.exit(1)

    qa_data = _load_qa_json(qa_json_path)
    qa_list: list[dict[str, Any]] = qa_data.get("qa_list", [])
    project_name: str = qa_data.get("project_name", "")

    answered_count = 0
    withdrawn_count = 0

    for item in qa_list:
        qa_id = item.get("id", "")
        if qa_id not in answers:
            continue

        answer = answers[qa_id]

        # 取下げキーワードの確認
        if any(kw in answer for kw in WITHDRAW_KEYWORDS):
            item["status"] = "取下げ"
            item["answer"] = answer
            withdrawn_count += 1
        else:
            item["status"] = "回答済み"
            item["answer"] = answer
            item["reflection_note"] = _generate_reflection_note(item, answer)
            answered_count += 1

    # qa_draft.jsonを更新
    qa_data["qa_list"] = qa_list
    qa_json_path.write_text(json.dumps(qa_data, ensure_ascii=False, indent=2), encoding="utf-8")

    # answers.md を生成
    total = len(qa_list)
    unanswered = [item for item in qa_list if item.get("status") == "未回答"]
    answered_items = [item for item in qa_list if item.get("status") == "回答済み"]
    withdrawn_items = [item for item in qa_list if item.get("status") == "取下げ"]

    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    lines: list[str] = [
        f"# 回答取込結果 - {project_name}",
        "",
        f"取込日時: {now_str}",
        f"回答済み: {answered_count}件 / 全{total}件（取下げ: {withdrawn_count}件、未回答: {len(unanswered)}件）",
        "",
        "---",
        "",
        "## 回答内容サマリ",
        "",
        "| No | 観点 | 優先度 | 質疑内容 | 回答 | 提案への示唆 |",
        "|----|------|--------|----------|------|------------|",
    ]

    for item in answered_items:
        qa_id = item.get("id", "")
        perspective = item.get("perspective", "")
        priority = item.get("priority", "")
        question = item.get("question", "").replace("|", "｜")[:50]
        answer = (item.get("answer") or "").replace("|", "｜")[:50]
        reflection = (item.get("reflection_note") or "").replace("|", "｜")[:50]
        lines.append(f"| {qa_id} | {perspective} | {priority} | {question} | {answer} | {reflection} |")

    if withdrawn_items:
        lines += [
            "",
            "## 取下げになった質疑",
            "",
            "| No | 質疑内容 |",
            "|----|----------|",
        ]
        for item in withdrawn_items:
            qa_id = item.get("id", "")
            question = item.get("question", "").replace("|", "｜")[:60]
            lines.append(f"| {qa_id} | {question} |")

    if unanswered:
        lines += [
            "",
            "## 未回答の質疑（提案書作成前に確認推奨）",
            "",
        ]
        for item in unanswered:
            priority = item.get("priority", "")
            qa_id = item.get("id", "")
            question = item.get("question", "")
            lines.append(f"- [{priority}] {qa_id}: {question}")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"回答取込完了: {output_path}")
    print(f"  回答済み: {answered_count}件, 取下げ: {withdrawn_count}件, 未回答: {len(unanswered)}件")


def _build_arg_parser() -> argparse.ArgumentParser:
    """コマンドライン引数パーサーを構築する。

    Returns:
        設定済みのArgumentParserオブジェクト。
    """
    parser = argparse.ArgumentParser(
        description="QA Excel ハンドラー",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    subparsers = parser.add_subparsers(dest="command", required=True)

    # write-excel サブコマンド
    p_write = subparsers.add_parser("write-excel", help="テンプレートを使ってQA一覧をExcelに出力する")
    p_write.add_argument("--template", required=True, type=Path, help="BoxからDLしたテンプレートExcelのパス")
    p_write.add_argument("--qa-json", required=True, type=Path, help="qa_draft.jsonのパス")
    p_write.add_argument("--output", required=True, type=Path, help="出力先Excelファイルのパス")

    # import-answers サブコマンド
    p_import = subparsers.add_parser("import-answers", help="回答済みExcel/Markdownを取り込む")
    p_import.add_argument("--answered", required=True, type=Path, help="顧客が回答済みのExcelまたはMarkdownのパス")
    p_import.add_argument("--qa-json", required=True, type=Path, help="qa_draft.jsonのパス")
    p_import.add_argument("--output", required=True, type=Path, help="answers.mdの出力先パス")

    return parser


def main() -> None:
    """エントリーポイント。コマンドライン引数を解析してサブコマンドを実行する。"""
    parser = _build_arg_parser()
    args = parser.parse_args()

    if args.command == "write-excel":
        cmd_write_excel(args.template, args.qa_json, args.output)
    elif args.command == "import-answers":
        cmd_import_answers(args.answered, args.qa_json, args.output)


if __name__ == "__main__":
    main()
